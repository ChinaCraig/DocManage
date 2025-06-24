from .base_preview import BasePreviewService
import os
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class ExcelPreviewService(BasePreviewService):
    """Excel文档预览服务"""
    
    def __init__(self):
        super().__init__()
        self.supported_formats = ['xlsx', 'xls']
    
    def extract_content(self, file_path, document_id=None):
        """提取Excel文档内容"""
        
        # 检查是否为MCP创建的虚拟文件
        if file_path and file_path.startswith('mcp_created/'):
            return self._extract_mcp_content(file_path, document_id)
        
        # 普通文件处理
        if not self.validate_file(file_path):
            raise FileNotFoundError(f"文件不存在或无法读取: {file_path}")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            content_data = {
                'sheets': [],
                'metadata': {}
            }
            
            if file_ext == '.xlsx':
                content_data = self._extract_xlsx_content(file_path)
            elif file_ext == '.xls':
                content_data = self._extract_xls_content(file_path)
            
            logger.info(f"✅ Excel内容提取成功: {file_path}")
            return content_data
            
        except Exception as e:
            logger.error(f"❌ Excel内容提取失败: {file_path}, 错误: {str(e)}")
            raise
    
    def _extract_mcp_content(self, file_path, document_id):
        """提取MCP创建的Excel文件内容"""
        try:
            logger.info(f"🔍 处理MCP Excel文件: {file_path}")
            
            # 导入模型
            from app.models.document_models import DocumentNode, DocumentContent
            from app import db
            
            if document_id:
                # 使用document_id查找
                document = db.session.query(DocumentNode).filter_by(
                    id=document_id,
                    is_deleted=False
                ).first()
            else:
                # 使用file_path查找
                document = db.session.query(DocumentNode).filter_by(
                    file_path=file_path,
                    is_deleted=False
                ).first()
                
            if not document:
                raise FileNotFoundError(f"MCP Excel文件不存在: {file_path}")
            
            # 检查虚拟文件是否存在于文件系统中
            import os
            full_virtual_path = os.path.join('uploads', file_path)
            
            if os.path.exists(full_virtual_path):
                # 文件存在于文件系统中，使用实际Excel文件处理
                logger.info(f"📁 虚拟Excel文件存在于文件系统: {full_virtual_path}")
                
                try:
                    file_ext = os.path.splitext(full_virtual_path)[1].lower()
                    if file_ext == '.xlsx':
                        return self._extract_xlsx_content(full_virtual_path)
                    elif file_ext == '.xls':
                        return self._extract_xls_content(full_virtual_path)
                    else:
                        raise ValueError(f"不支持的Excel格式: {file_ext}")
                        
                except Exception as e:
                    logger.warning(f"处理虚拟Excel文件失败，降级到文本模式: {e}")
                    # 降级到文本内容处理
                    pass
            
            # 从数据库获取文本内容（降级处理）
            logger.info(f"💾 从数据库读取MCP Excel文件内容: {file_path}")
            
            content_record = db.session.query(DocumentContent).filter_by(
                document_id=document.id
            ).first()
            
            if content_record and content_record.content_text:
                # 将文本内容转换为Excel预览格式
                text_content = content_record.content_text
                
                # 模拟Excel工作表结构
                content_data = {
                    'sheets': [{
                        'name': '生成内容',
                        'max_row': len(text_content.split('\n')),
                        'max_column': 1,
                        'data': [[line] for line in text_content.split('\n') if line.strip()]
                    }],
                    'metadata': {
                        'sheet_count': 1,
                        'sheet_names': ['生成内容'],
                        'source': 'mcp_created',
                        'format': 'xlsx',
                        'note': '📄 Excel文档已生成，可通过下载功能获取完整表格文件'
                    }
                }
                
                logger.info(f"✅ MCP Excel文件内容提取成功: {file_path}")
                return content_data
            else:
                # 返回空内容
                return {
                    'sheets': [{
                        'name': 'Empty',
                        'max_row': 1,
                        'max_column': 1,
                        'data': [['[无内容]']]
                    }],
                    'metadata': {
                        'sheet_count': 1,
                        'sheet_names': ['Empty'],
                        'source': 'mcp_created',
                        'format': 'xlsx',
                        'error': '无法获取文件内容'
                    }
                }
            
        except Exception as e:
            logger.error(f"❌ MCP Excel文件内容提取失败: {file_path}, 错误: {str(e)}")
            raise
    
    def _extract_xlsx_content(self, file_path):
        """提取XLSX文档内容"""
        content_data = {
            'sheets': [],
            'metadata': {}
        }
        
        # 使用openpyxl读取文件
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        
        # 获取工作表信息
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # 获取工作表数据
            sheet_data = {
                'name': sheet_name,
                'max_row': worksheet.max_row,
                'max_column': worksheet.max_column,
                'data': []
            }
            
            # 读取前100行数据作为预览
            max_preview_rows = min(100, worksheet.max_row)
            
            for row in worksheet.iter_rows(min_row=1, max_row=max_preview_rows, values_only=True):
                # 过滤空行
                if any(cell is not None for cell in row):
                    # 转换为字符串，处理None值
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    sheet_data['data'].append(row_data)
            
            content_data['sheets'].append(sheet_data)
        
        workbook.close()
        
        # 获取元数据
        content_data['metadata'] = {
            'sheet_count': len(workbook.sheetnames),
            'sheet_names': workbook.sheetnames
        }
        
        return content_data
    
    def _extract_xls_content(self, file_path):
        """提取XLS文档内容"""
        content_data = {
            'sheets': [],
            'metadata': {}
        }
        
        # 使用pandas读取文件
        excel_file = pd.ExcelFile(file_path)
        
        for sheet_name in excel_file.sheet_names:
            # 读取工作表数据
            df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=100)
            
            # 转换为列表格式
            sheet_data = {
                'name': sheet_name,
                'max_row': len(df) + 1,  # +1 for header
                'max_column': len(df.columns),
                'data': []
            }
            
            # 添加列标题
            if not df.empty:
                sheet_data['data'].append(df.columns.tolist())
                
                # 添加数据行
                for _, row in df.iterrows():
                    row_data = [str(val) if pd.notna(val) else '' for val in row.values]
                    sheet_data['data'].append(row_data)
            
            content_data['sheets'].append(sheet_data)
        
        # 获取元数据
        content_data['metadata'] = {
            'sheet_count': len(excel_file.sheet_names),
            'sheet_names': excel_file.sheet_names
        }
        
        excel_file.close()
        
        return content_data
    
    def get_metadata(self, file_path):
        """获取Excel文档元数据"""
        if not self.validate_file(file_path):
            raise FileNotFoundError(f"文件不存在或无法读取: {file_path}")
        
        try:
            metadata = {
                'file_size': self.get_file_size(file_path),
                'file_size_formatted': self.format_file_size(self.get_file_size(file_path)),
                'file_type': 'Excel Spreadsheet',
                'last_modified': datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat()
            }
            
            file_ext = os.path.splitext(file_path)[1].lower()
            
            try:
                if file_ext == '.xlsx':
                    # 使用openpyxl获取详细信息
                    workbook = load_workbook(file_path, read_only=True)
                    
                    # 获取工作表信息
                    sheet_info = []
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        sheet_info.append({
                            'name': sheet_name,
                            'max_row': worksheet.max_row,
                            'max_column': worksheet.max_column
                        })
                    
                    metadata.update({
                        'sheet_count': len(workbook.sheetnames),
                        'sheet_names': workbook.sheetnames,
                        'sheets_info': sheet_info
                    })
                    
                    # 尝试获取文档属性
                    if hasattr(workbook, 'properties') and workbook.properties:
                        props = workbook.properties
                        metadata.update({
                            'title': props.title or '',
                            'author': props.creator or '',
                            'subject': props.subject or '',
                            'keywords': props.keywords or '',
                            'comments': props.description or '',
                            'created': props.created.isoformat() if props.created else '',
                            'modified': props.modified.isoformat() if props.modified else ''
                        })
                    
                    workbook.close()
                    
                elif file_ext == '.xls':
                    # 使用pandas获取基本信息
                    excel_file = pd.ExcelFile(file_path)
                    
                    sheet_info = []
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=0)
                        sheet_info.append({
                            'name': sheet_name,
                            'columns': list(df.columns)
                        })
                    
                    metadata.update({
                        'sheet_count': len(excel_file.sheet_names),
                        'sheet_names': excel_file.sheet_names,
                        'sheets_info': sheet_info,
                        'format_note': 'XLS格式，建议转换为XLSX以获取更多信息'
                    })
                    
                    excel_file.close()
                    
            except Exception as e:
                logger.warning(f"无法提取Excel详细元数据: {file_path}, 错误: {str(e)}")
                metadata['error'] = str(e)
            
            return metadata
            
        except Exception as e:
            logger.error(f"获取Excel元数据失败: {file_path}, 错误: {str(e)}")
            return {
                'file_size': self.get_file_size(file_path),
                'file_size_formatted': self.format_file_size(self.get_file_size(file_path)),
                'file_type': 'Excel Spreadsheet',
                'last_modified': datetime.fromtimestamp(os.path.getmtime(file_path)).isoformat(),
                'error': str(e)
            }
    
    def generate_thumbnail(self, file_path, output_path=None, size=(200, 200)):
        """生成Excel文档缩略图"""
        # Excel文档缩略图生成需要特殊工具
        # 目前返回None，表示不支持
        logger.info(f"Excel文档缩略图生成暂不支持: {file_path}")
        return None